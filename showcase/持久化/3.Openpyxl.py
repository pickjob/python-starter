#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl

if __name__ == '__main__':
    # workbook
    wb = openpyxl.Workbook()
    # 读取Excel文档
    wb = openpyxl.load_workbook('document.xlsx', read_only=True)
    # 保存Excel文档
    wb.save('balances.xlsx')

    # worksheet
    ws = wb.create_sheet("Mysheet")
    ws = wb["New Title"]
    for sheet in wb:
        print(sheet.title)

    # cell、col、row value
    c = ws['A4']
    c = ws.cell(row=4, column=2, value=10)
    cell_range = ws['A1':'C2']
    colC = ws['C']
    col_range = ws['C:D']
    row10 = ws[10]
    row_range = ws[5:10]

    ws['A4'] = 4

    # iterator
    for row in ws.values:
        for value in row:
            print(value)

    # iterator rows
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                print(cell.value)

    # iterator cols
    for col in ws.iter_cols():
        for cell in col:
            if cell.value is not None:
                print(cell.value)
